from __future__ import annotations

import csv
import os
import secrets
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO, StringIO

from flask import Flask, Response, abort, jsonify, redirect, render_template, request, session, url_for, send_file
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from config import Config
from detection_schema import DETECTION_CLASS_NAMES, class_id_from_payload, class_name_from_id
from monitoring_store import store
from water_quality_rules import classify_water_quality


app = Flask(__name__)
app.config.from_object(Config)
app.config.setdefault("PERMANENT_SESSION_LIFETIME", timedelta(hours=8))


@app.after_request
def set_security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; "
        "script-src 'self' https://cdn.jsdelivr.net 'unsafe-inline'; "
        "style-src 'self' https://cdn.jsdelivr.net https://fonts.googleapis.com 'unsafe-inline'; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data: https:; "
        "connect-src 'self'",
    )
    if app.config.get("SESSION_COOKIE_SECURE"):
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return response


@dataclass(frozen=True)
class User:
    email: str
    password_hash: str
    role: str


def _build_users() -> dict[str, User]:
    """Build the login table with hashed passwords sourced from env/config.

    Passwords are never stored or compared in plaintext. Set ADMIN_PASSWORD,
    RESEARCHER_PASSWORD, and VIEWER_PASSWORD env vars for production.
    """
    accounts = [
        (Config.ADMIN_EMAIL, Config.ADMIN_PASSWORD, "Admin"),
        (Config.RESEARCHER_EMAIL, Config.RESEARCHER_PASSWORD, "Researcher"),
        (Config.VIEWER_EMAIL, Config.VIEWER_PASSWORD, "Viewer"),
    ]
    table: dict[str, User] = {}
    for email, password, role in accounts:
        email = email.strip().lower()
        table[email] = User(email, generate_password_hash(password), role)
    return table


USERS = _build_users()

# --- CSRF protection (dependency-free, for session-cookie form POSTs) ---


def get_csrf_token() -> str:
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token


def csrf_ok() -> bool:
    expected = session.get("_csrf_token")
    submitted = request.form.get("_csrf_token", "")
    return bool(expected) and secrets.compare_digest(expected, submitted)


@app.context_processor
def inject_csrf_token():
    return {"csrf_token": get_csrf_token}


# --- Simple in-memory login rate limiter (per client IP) ---
_login_attempts: dict[str, list[float]] = defaultdict(list)


def login_rate_limited() -> bool:
    ip = (request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0]).strip()
    window = Config.LOGIN_RATE_WINDOW
    limit = Config.LOGIN_RATE_LIMIT
    now = time.time()
    attempts = [t for t in _login_attempts[ip] if now - t < window]
    _login_attempts[ip] = attempts
    if len(attempts) >= limit:
        return True
    _login_attempts[ip].append(now)
    return False


ROLE_DESCRIPTIONS = {
    "Admin": "Full access to dashboard, reports, alerts, and users.",
    "Researcher": "Access to monitoring, historical analysis, and exports.",
    "Viewer": "Read-only access to live monitoring and logs.",
}


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_email"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


def current_user():
    if not session.get("user_email"):
        return None
    return {
        "email": session.get("user_email"),
        "role": session.get("user_role", "Viewer"),
    }


def require_role(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            user_role = session.get("user_role")
            if not session.get("user_email"):
                return redirect(url_for("login"))
            if roles and user_role not in roles:
                return abort(403)
            return view(*args, **kwargs)

        return wrapped

    return decorator


def parse_filters(args):
    query = args.get("query", "").strip().lower()
    return {
        "period": args.get("period", "today"),
        "date_from": args.get("date_from", ""),
        "date_to": args.get("date_to", ""),
        "object_type": args.get("object_type", "all"),
        "status": args.get("status", "all"),
        "query": query,
        "search": query,
        "dataset": args.get("dataset", "water_quality"),
    }


def require_ingest_token():
    expected = app.config.get("INGEST_API_KEY")
    if not expected:
        return True
    token = request.headers.get("X-INGEST-TOKEN") or request.args.get("token", "")
    return token == expected


def parse_date(value: str):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return None


def report_summary(water_rows, detection_rows):
    return {
        "water_samples": len(water_rows),
        "detections": len(detection_rows),
        "good_status": sum(1 for row in water_rows if row["status"] == "Good"),
        "moderate_status": sum(1 for row in water_rows if row["status"] == "Moderate"),
        "poor_status": sum(1 for row in water_rows if row["status"] == "Poor"),
        "total_objects": sum(row["total_objects"] for row in detection_rows),
    }


def csv_response(rows, filename, fieldnames):
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    response = Response(buffer.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def xlsx_response(rows, filename, headers):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"
    worksheet.append(list(headers.values()))
    for row in rows:
        worksheet.append([row[key] for key in headers.keys()])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def pdf_response(rows, filename, headers):
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [Paragraph("Smart AI Lake Monitoring Report", styles["Title"]), Spacer(1, 12)]
    table_data = [list(headers.values())]
    for row in rows:
        table_data.append([str(row[key]) for key in headers.keys()])
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(table)
    document.build(story)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/pdf")


def combined_csv_response(report, filename):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Smart AI Lake Cleaning Robot Monitoring System"])
    writer.writerow([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"])
    writer.writerow([])
    writer.writerow(["Summary"])
    for key, value in report_summary(report["water_logs"], report["detections"]).items():
        writer.writerow([key, value])
    writer.writerow([])
    writer.writerow(["Detection Logs"])
    writer.writerow(["Timestamp", "Class ID", "Class Name", "Bottle Count", "Debris Count", "Total Objects", "Confidence"])
    for row in report["detections"]:
        writer.writerow([row["timestamp"], row["class_id"], row["class_name"], row["bottle_count"], row["debris_count"], row["total_objects"], row["confidence_score"]])
    writer.writerow([])
    writer.writerow(["Water Quality Logs"])
    writer.writerow(["Timestamp", "TDS", "Turbidity", "Temperature", "Status"])
    for row in report["water_logs"]:
        writer.writerow([row["timestamp"], row["tds"], row["turbidity"], row["temperature"], row["status"]])

    response = Response(buffer.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def combined_xlsx_response(report, filename):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    for key, value in report_summary(report["water_logs"], report["detections"]).items():
        summary_sheet.append([key, value])

    detection_sheet = workbook.create_sheet("Detections")
    detection_sheet.append(["Timestamp", "Class ID", "Class Name", "Bottle Count", "Debris Count", "Total Objects", "Confidence"])
    for row in report["detections"]:
        detection_sheet.append([row["timestamp"], row["class_id"], row["class_name"], row["bottle_count"], row["debris_count"], row["total_objects"], row["confidence_score"]])

    water_sheet = workbook.create_sheet("Water Quality")
    water_sheet.append(["Timestamp", "TDS", "Turbidity", "Temperature", "Status"])
    for row in report["water_logs"]:
        water_sheet.append([row["timestamp"], row["tds"], row["turbidity"], row["temperature"], row["status"]])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def combined_pdf_response(report, filename):
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [Paragraph("Smart AI Lake Monitoring Report", styles["Title"]), Spacer(1, 12)]

    summary_data = [["Metric", "Value"]]
    summary_data.extend([[key, str(value)] for key, value in report_summary(report["water_logs"], report["detections"]).items()])
    summary_table = Table(summary_data, repeatRows=1)
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ]
        )
    )
    story.extend([Paragraph("Summary", styles["Heading2"]), summary_table, Spacer(1, 12)])

    detection_data = [["Timestamp", "Class", "Bottle", "Debris", "Total", "Confidence"]]
    detection_data.extend(
        [[row["timestamp"], f"{row['class_id']}: {row['class_name']}", str(row["bottle_count"]), str(row["debris_count"]), str(row["total_objects"]), f"{row['confidence_score']:.2f}"] for row in report["detections"]]
    )
    detection_table = Table(detection_data, repeatRows=1)
    detection_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#102033")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ]
        )
    )
    story.extend([Paragraph("Detection Logs", styles["Heading2"]), detection_table, Spacer(1, 12)])

    water_data = [["Timestamp", "TDS", "Turbidity", "Temperature", "Status"]]
    water_data.extend([[row["timestamp"], str(row["tds"]), str(row["turbidity"]), str(row["temperature"]), row["status"]] for row in report["water_logs"]])
    water_table = Table(water_data, repeatRows=1)
    water_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ]
        )
    )
    story.extend([Paragraph("Water Quality Logs", styles["Heading2"]), water_table])

    document.build(story)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/pdf")


@app.route("/")
def root():
    if session.get("user_email"):
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        if login_rate_limited():
            error = "Too many login attempts. Please wait a few minutes and try again."
            return render_template("login.html", error=error, app_name=app.config["APP_NAME"]), 429
        if not csrf_ok():
            error = "Security token expired. Please try again."
            return render_template("login.html", error=error, app_name=app.config["APP_NAME"]), 400

        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = USERS.get(email)
        if user and check_password_hash(user.password_hash, password):
            session.clear()
            session.permanent = True
            session["user_email"] = user.email
            session["user_role"] = user.role
            return redirect(url_for("dashboard"))
        error = "Invalid email or password."

    return render_template("login.html", error=error, app_name=app.config["APP_NAME"])


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    message = None
    if request.method == "POST":
        if not csrf_ok():
            return render_template("forgot_password.html", message="Security token expired. Please try again.", app_name=app.config["APP_NAME"]), 400
        email = request.form.get("email", "").strip().lower()
        if email in USERS:
            message = f"Password reset request prepared for {email}."
        else:
            message = "If the email exists, a reset link would be sent in production."
    return render_template("forgot_password.html", message=message, app_name=app.config["APP_NAME"])


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    return render_template(
        "dashboard.html",
        app_name=app.config["APP_NAME"],
        user=current_user(),
    )


@app.route("/reports")
@login_required
def reports():
    filters = parse_filters(request.args)
    payload = store.report_payload(filters)
    return render_template(
        "reports.html",
        app_name=app.config["APP_NAME"],
        user=current_user(),
        filters=filters,
        report=payload,
        water_quality_statuses=["all", "Good", "Moderate", "Poor"],
        detection_types=["all", *DETECTION_CLASS_NAMES],
        summary=report_summary(payload["water_logs"], payload["detections"]),
        query_params=request.args.to_dict(flat=True),
    )


@app.route("/users")
@login_required
@require_role("Admin")
def users_page():
    return render_template(
        "users.html",
        app_name=app.config["APP_NAME"],
        user=current_user(),
        users=store.users(),
        role_descriptions=ROLE_DESCRIPTIONS,
    )


@app.route("/alerts")
@login_required
def alerts_page():
    return render_template("alerts.html", app_name=app.config["APP_NAME"], user=current_user(), alerts=store.alerts())


@app.route("/api/overview")
@login_required
def api_overview():
    return jsonify(store.overview())


@app.route("/api/ingest/water", methods=["POST"])
def ingest_water():
    if not require_ingest_token():
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    return jsonify({"ok": True, "water": store.record_water_sample(payload)})


@app.route("/api/ingest/detection", methods=["POST"])
def ingest_detection():
    if not require_ingest_token():
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    return jsonify({"ok": True, "detection": store.record_detection_sample(payload)})


@app.route("/api/ingest/frame", methods=["POST"])
def ingest_frame():
    if not require_ingest_token():
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    payload = request.form.to_dict()
    image_file = request.files.get("image")
    if image_file is None:
        payload.update(request.get_json(silent=True) or {})
    return jsonify({"ok": True, "frame": store.record_frame(payload, image_file=image_file)})


@app.route("/api/system-status")
@login_required
def api_system_status():
    payload = store.overview()["status"]
    payload.update(store.stream_status())
    return jsonify(payload)


@app.route("/api/water-quality/current")
@login_required
def api_water_quality_current():
    return jsonify(store.latest_water_quality())


@app.route("/api/charts/<period>")
@login_required
def api_charts(period: str):
    return jsonify(store.trend(period))


@app.route("/api/detections")
@login_required
def api_detections():
    return jsonify(store.detections(parse_filters(request.args)))


@app.route("/api/water-quality/logs")
@login_required
def api_water_quality_logs():
    return jsonify(store.water_logs(parse_filters(request.args)))


@app.route("/api/search")
@login_required
def api_search():
    filters = parse_filters(request.args)
    return jsonify(store.search(filters))


@app.route("/reports/export/<fmt>")
@login_required
def export_report(fmt: str):
    filters = parse_filters(request.args)
    report = store.report_payload(filters)
    dataset = request.args.get("dataset", "combined")

    if dataset == "detections":
        if fmt == "csv":
            return csv_response(
                report["detections"],
                "detection-report.csv",
                ["timestamp", "class_id", "class_name", "bottle_count", "debris_count", "total_objects", "confidence_score"],
            )
        if fmt == "xlsx":
            return xlsx_response(
                report["detections"],
                "detection-report.xlsx",
                {
                    "timestamp": "Timestamp",
                    "class_id": "Class ID",
                    "class_name": "Class Name",
                    "bottle_count": "Bottle Count",
                    "debris_count": "Debris Count",
                    "total_objects": "Total Objects",
                    "confidence_score": "Confidence",
                },
            )
        if fmt == "pdf":
            return pdf_response(
                report["detections"],
                "detection-report.pdf",
                {
                    "timestamp": "Timestamp",
                    "class_id": "Class ID",
                    "class_name": "Class Name",
                    "bottle_count": "Bottle Count",
                    "debris_count": "Debris Count",
                    "total_objects": "Total Objects",
                    "confidence_score": "Confidence",
                },
            )
    elif dataset == "water_quality":
        if fmt == "csv":
            return csv_response(
                report["water_logs"],
                "water-quality-report.csv",
                ["timestamp", "tds", "turbidity", "temperature", "status"],
            )
        if fmt == "xlsx":
            return xlsx_response(
                report["water_logs"],
                "water-quality-report.xlsx",
                {
                    "timestamp": "Timestamp",
                    "tds": "TDS",
                    "turbidity": "Turbidity",
                    "temperature": "Temperature",
                    "status": "Status",
                },
            )
        if fmt == "pdf":
            return pdf_response(
                report["water_logs"],
                "water-quality-report.pdf",
                {
                    "timestamp": "Timestamp",
                    "tds": "TDS",
                    "turbidity": "Turbidity",
                    "temperature": "Temperature",
                    "status": "Status",
                },
            )
    else:
        if fmt == "csv":
            return combined_csv_response(report, "monitoring-report.csv")
        if fmt == "xlsx":
            return combined_xlsx_response(report, "monitoring-report.xlsx")
        if fmt == "pdf":
            return combined_pdf_response(report, "monitoring-report.pdf")

    abort(404)


@app.route("/api/stream/status")
@login_required
def stream_status():
    return jsonify(store.stream_status())


@app.route("/api/stream/snapshot")
@login_required
def stream_snapshot():
    return jsonify(store.latest_frame())


@app.route("/api/users")
@login_required
@require_role("Admin")
def api_users():
    return jsonify(store.users())


@app.route("/api/roles")
@login_required
def api_roles():
    return jsonify({"roles": list(ROLE_DESCRIPTIONS.keys())})


@app.errorhandler(403)
def forbidden(_):
    return render_template("forbidden.html", app_name=app.config["APP_NAME"], user=current_user()), 403


@app.errorhandler(404)
def page_not_found(_):
    if session.get("user_email"):
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


if __name__ == "__main__":
    # Only enable debug for local development
    debug_mode = app.config.get("DEBUG", False)
    app.run(debug=debug_mode, host="0.0.0.0", port=int(os.getenv("PORT", 5000)))
