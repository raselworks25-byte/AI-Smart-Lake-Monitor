from __future__ import annotations
import json
import base64
import csv
import os
from dataclasses import dataclass
from datetime import datetime, timedelta
from io import BytesIO, StringIO
from typing import Any, Iterable

from config import Config
from detection_schema import DETECTION_CLASS_NAMES, class_id_from_payload, class_name_from_id, class_name_from_payload
from water_quality_rules import classify_water_quality
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

try:
    import firebase_admin
    from firebase_admin import credentials, db, storage
except Exception:  # pragma: no cover - firebase is optional during local development
    firebase_admin = None
    credentials = None
    db = None
    storage = None


def _utc(value: datetime) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S UTC")


def _parse_date(value: str | None) -> datetime | None:
    if not value:
        return None
    formats = ["%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"]
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _window_start(period: str | None) -> datetime | None:
    if not period or period == "all":
        return None
    now = datetime.utcnow()
    if period == "today":
        return datetime(now.year, now.month, now.day)
    if period == "week":
        return now - timedelta(days=7)
    if period == "month":
        return now - timedelta(days=30)
    if period == "year":
        return now - timedelta(days=365)
    return None


def _matches_window(timestamp: datetime, filters: dict[str, str]) -> bool:
    start = _window_start(filters.get("period"))
    if start and timestamp < start:
        return False

    exact_date = _parse_date(filters.get("date"))
    if exact_date and timestamp.date() != exact_date.date():
        return False

    start_date = _parse_date(filters.get("start"))
    end_date = _parse_date(filters.get("end"))
    if start_date and timestamp < start_date:
        return False
    if end_date and timestamp > end_date + timedelta(days=1):
        return False

    return True


def _normalize(value: str | None) -> str:
    return (value or "").strip().lower()


def _parse_timestamp(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    formats = ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S UTC", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%d"]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


@dataclass(frozen=True)
class UserAccount:
    email: str
    password: str
    role: str
    name: str
    active: bool = True
    password_hash: str = ""


class MonitoringStore:
    def __init__(self) -> None:
        now = datetime.utcnow().replace(microsecond=0)

        # Built-in accounts, seeded with hashed passwords from config/env.
        # Admins can add more at runtime from the Users page (persisted to
        # Firebase when enabled, so they survive restarts).
        self.user_accounts = [
            UserAccount("admin@lakewatch.local", "", "Admin", "System Admin",
                        password_hash=generate_password_hash(Config.ADMIN_PASSWORD)),
            UserAccount("researcher@lakewatch.local", "", "Researcher", "Research Lead",
                        password_hash=generate_password_hash(Config.RESEARCHER_PASSWORD)),
            UserAccount("viewer@lakewatch.local", "", "Viewer", "Field Viewer",
                        password_hash=generate_password_hash(Config.VIEWER_PASSWORD)),
        ]

        self.thresholds = {"tds": 300, "turbidity": 10, "temperature": 35}
        self.firebase_enabled = False
        self.firebase_root = None
        self.latest_frame_data = {"snapshot_url": "/static/img/placeholder-detection.svg", "timestamp": _utc(now)}

        self.system_snapshot = {
            "raspberry_pi": "Offline",
            "esp32": "Offline",
            "camera": "Offline",
            "cloud": "Local",
            "last_update": _utc(now),
        }

        # Alerts start empty; they are populated only from real events / Firebase.
        self.alert_log: list[dict[str, Any]] = []

        # Logs start empty. They fill only with REAL data ingested from the
        # Raspberry Pi via /api/ingest/water and /api/ingest/detection.
        self.detection_log: list[dict[str, Any]] = []
        self.water_quality_log: list[dict[str, Any]] = []

        # Tracks when the last real reading arrived, so device status reflects
        # actual connectivity instead of a hardcoded "Online".
        self.last_ingest_at: datetime | None = None

        # Cumulative detection totals -- grow over time, never reset.
        # Persisted to Firebase so they survive restarts.
        self.total_bottles = 0
        self.total_debris = 0
        self._prev_bottle = 0
        self._prev_debris = 0

        self._bootstrap_firebase()
        self._load_persisted_users()
        self._load_detection_totals()

    def _bootstrap_firebase(self) -> None:
        # Firebase is optional. If anything goes wrong (missing credentials,
        # bad URL, no network), log a warning and fall back to in-memory
        # storage instead of crashing the whole app on startup.
        if not firebase_admin or not Config.FIREBASE_DATABASE_URL:
            return

        try:
            try:
                firebase_admin.get_app()
            except ValueError:
                credentials_path = Config.FIREBASE_CREDENTIALS_PATH
                if credentials_path and os.path.exists(credentials_path):
                    credential = credentials.Certificate(credentials_path)
                else:
                    # No service-account file present. Try Application Default
                    # Credentials; if they aren't configured this will raise,
                    # and we fall back to in-memory below.
                    credential = credentials.ApplicationDefault()

                options: dict[str, Any] = {"databaseURL": Config.FIREBASE_DATABASE_URL}
                if Config.FIREBASE_STORAGE_BUCKET:
                    options["storageBucket"] = Config.FIREBASE_STORAGE_BUCKET
                firebase_admin.initialize_app(credential, options)

            # Validating the connection also happens here, inside the guard.
            self.firebase_root = db.reference("monitoring")
            self.firebase_enabled = True
            print("[monitoring_store] Firebase connected.")
        except Exception as exc:  # noqa: BLE001 - any failure -> in-memory mode
            self.firebase_enabled = False
            self.firebase_root = None
            print(f"[monitoring_store] Firebase unavailable, using in-memory storage: {exc}")

    def _use_firebase(self) -> bool:
        return self.firebase_enabled and self.firebase_root is not None

    def _root_child(self, path: str):
        if not self._use_firebase():
            return None
        return self.firebase_root.child(path)

    def _load_detection_totals(self) -> None:
        """Load cumulative detection totals from Firebase so they survive restarts."""
        if not self._use_firebase():
            return
        try:
            data = self._read_value("detection/totals", {}) or {}
            self.total_bottles = int(data.get("total_bottles", 0) or 0)
            self.total_debris = int(data.get("total_debris", 0) or 0)
        except Exception:
            pass

    def _read_collection(self, path: str) -> list[dict[str, Any]]:
        ref = self._root_child(path)
        if ref is None:
            return []
        raw = ref.get() or {}
        if isinstance(raw, list):
            rows = [item for item in raw if isinstance(item, dict)]
        elif isinstance(raw, dict):
            rows = [item for item in raw.values() if isinstance(item, dict)]
        else:
            rows = []
        rows.sort(key=lambda row: _parse_timestamp(row.get("timestamp")) or datetime.min, reverse=True)
        return rows

    def _read_value(self, path: str, default: dict[str, Any]) -> dict[str, Any]:
        ref = self._root_child(path)
        if ref is None:
            return default
        raw = ref.get()
        return raw if isinstance(raw, dict) and raw else default

    def _write_value(self, path: str, payload: dict[str, Any]) -> dict[str, Any]:
        ref = self._root_child(path)
        if ref is not None:
            ref.set(payload)
        return payload

    def _push_record(self, path: str, payload: dict[str, Any]) -> dict[str, Any]:
        ref = self._root_child(path)
        if ref is not None:
            ref.push(payload)
        return payload

    def _placeholder_frame(self) -> dict[str, Any]:
        return {"snapshot_url": "/static/img/placeholder-detection.svg", "timestamp": _utc(datetime.utcnow())}

    def _quality_status(self, tds: int, turbidity: float, temperature: float) -> str:
        return classify_water_quality(tds, turbidity)

    def _serialise_detection(self, row: dict[str, Any]) -> dict[str, Any]:
        timestamp = _parse_timestamp(row.get("timestamp"))
        class_id = class_id_from_payload(row)
        class_name = row.get("class_name") or row.get("object_type") or class_name_from_id(class_id)
        return {
            "timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S") if timestamp else str(row.get("timestamp", "")),
            "class_id": class_id,
            "class_name": class_name,
            "object_type": class_name,
            "bottle_count": row["bottle_count"],
            "debris_count": row.get("debris_count", 0),
            "total_objects": row["total_objects"],
            "confidence_score": row.get("confidence_score", row.get("confidence")),
            "image_url": row.get("image_url", row.get("image", "/static/img/placeholder-detection.svg")),
        }

    def _serialise_water(self, row: dict[str, Any]) -> dict[str, Any]:
        timestamp = _parse_timestamp(row.get("timestamp"))
        return {
            "timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S") if timestamp else str(row.get("timestamp", "")),
            "tds": row["tds"],
            "turbidity": row["turbidity"],
            "temperature": row["temperature"],
            "status": row["status"],
        }

    def _apply_filters(self, rows: Iterable[dict[str, Any]], filters: dict[str, str], kind: str) -> list[dict[str, Any]]:
        object_type = _normalize(filters.get("object_type"))
        status = _normalize(filters.get("status"))
        search = _normalize(filters.get("search") or filters.get("query"))
        # "all" (or empty) means: do not filter on that field.
        if object_type in ("", "all"):
            object_type = ""
        if status in ("", "all"):
            status = ""

        results: list[dict[str, Any]] = []
        for row in rows:
            timestamp = _parse_timestamp(row.get("timestamp"))
            if not timestamp:
                continue
            if not _matches_window(timestamp, filters):
                continue

            if kind == "detection":
                if object_type and object_type not in _normalize(row["object_type"]):
                    continue
                if search and search not in _normalize(row["object_type"]):
                    continue
                results.append(self._serialise_detection(row))
            else:
                if status and status not in _normalize(row["status"]):
                    continue
                if search and search not in _normalize(row["status"]):
                    continue
                results.append(self._serialise_water(row))

        return results

    def overview(self) -> dict[str, Any]:
        water = self.latest_water_quality()
        detection = self.detection_summary()
        status = dict(self.system_snapshot)
        stream_state = self.stream_status()
        recent_ingest = (
            self.last_ingest_at is not None
            and (datetime.utcnow() - self.last_ingest_at) <= timedelta(minutes=2)
        )
        status["raspberry_pi"] = "Online" if recent_ingest else "Offline"
        status["esp32"] = "Online" if recent_ingest else "Offline"
        status["camera"] = "Streaming" if stream_state["streaming"] else "Offline"
        status["cloud"] = "Connected" if self._use_firebase() else "Local"
        status["last_update"] = _utc(self.last_ingest_at) if self.last_ingest_at else self.system_snapshot["last_update"]
        return {
            "status": status,
            "water_quality": water,
            "detection_summary": detection,
            "alerts": self.alerts(),
        }

    def latest_water_quality(self) -> dict[str, Any]:
        rows = self._read_collection("water_quality_logs") if self._use_firebase() else self.water_quality_log
        if not rows:
            return {"tds": 0, "turbidity": 0, "temperature": 0, "status": "No Data", "thresholds": self.thresholds, "last_update": _utc(datetime.utcnow())}
        latest = rows[0]
        latest_timestamp = _parse_timestamp(latest.get("timestamp"))
        return {
            "tds": latest["tds"],
            "turbidity": latest["turbidity"],
            "temperature": latest["temperature"],
            "status": latest["status"],
            "thresholds": self.thresholds,
            "last_update": latest_timestamp.strftime("%Y-%m-%d %H:%M:%S") if latest_timestamp else _utc(datetime.utcnow()),
        }

    def detection_summary(self) -> dict[str, Any]:
        rows = self._read_collection("detection_logs") if self._use_firebase() else self.detection_log
        latest_rows = rows[:10]
        latest = latest_rows[0] if latest_rows else {}
        # Cumulative totals -- grow over time, never reset to 0.
        plastic_bottle = self.total_bottles
        debris = self.total_debris
        total_waste = self.total_bottles + self.total_debris
        return {
            "plastic_bottle": plastic_bottle,
            "debris": debris,
            "total_waste": total_waste,
            "latest_confidence": latest_rows[0].get("confidence_score", latest_rows[0].get("confidence", 0)) if latest_rows else 0,
            "latest_class_name": class_name_from_payload(latest_rows[0]) if latest_rows else "No Data",
        }

    def alerts(self) -> list[dict[str, Any]]:
        if self._use_firebase():
            alerts = self._read_collection("alerts")
            if alerts:
                return alerts
            water = self.latest_water_quality()
            frame = self.latest_frame()
            generated = [
                {"level": "warning", "message": f"Turbidity is {water['turbidity']} NTU. Review threshold handling.", "time": "Just now"},
                {"level": "info", "message": f"Last camera frame received at {frame['timestamp']}.", "time": "2 min ago"},
            ]
            if water["status"] == "Poor":
                generated.insert(0, {"level": "danger", "message": "Water quality dropped to Poor. Immediate attention required.", "time": "Just now"})
            return generated
        return list(self.alert_log)

    def detections(self, filters: dict[str, str]) -> list[dict[str, Any]]:
        rows = self._read_collection("detection_logs") if self._use_firebase() else self.detection_log
        return self._apply_filters(rows, filters, kind="detection")

    def water_logs(self, filters: dict[str, str]) -> list[dict[str, Any]]:
        rows = self._read_collection("water_quality_logs") if self._use_firebase() else self.water_quality_log
        return self._apply_filters(rows, filters, kind="water")

    def trend(self, period: str | None) -> dict[str, list[Any]]:
        filters = {"period": period or "today"}
        rows = self.water_logs(filters)
        rows.reverse()

        labels = [row["timestamp"].split(" ")[0] if period in {"week", "month", "year"} else row["timestamp"][11:16] for row in rows[-8:]]
        return {
            "labels": labels,
            "tds": [row["tds"] for row in rows[-8:]],
            "turbidity": [row["turbidity"] for row in rows[-8:]],
            "temperature": [row["temperature"] for row in rows[-8:]],
            "waste": [row["total_objects"] if "total_objects" in row else 0 for row in self.detections({"period": period or "today"})[:8]],
        }

    def stream_status(self) -> dict[str, Any]:
        frame = self.latest_frame()
        frame_timestamp = _parse_timestamp(frame.get("timestamp"))
        is_live = bool(frame.get("snapshot_url")) and frame.get("snapshot_url") != "/static/img/placeholder-detection.svg"
        if frame_timestamp:
            is_live = is_live and (datetime.utcnow() - frame_timestamp) <= timedelta(minutes=2)
        return {
            "streaming": is_live,
            "fps": 24,
            "snapshot_ready": bool(frame.get("snapshot_url")),
            "timestamp": _utc(datetime.utcnow()),
            "snapshot_url": frame.get("snapshot_url", "/static/img/placeholder-detection.svg"),
        }

    def latest_frame(self) -> dict[str, Any]:
        frame = dict(self.latest_frame_data)
        return {
            "snapshot_url": frame.get("snapshot_url", "/static/img/placeholder-detection.svg"),
            "timestamp": frame.get("timestamp", _utc(datetime.utcnow())),
            "device_id": frame.get("device_id", "raspberry-pi"),
        }

    def record_water_sample(self, payload: dict[str, Any]) -> dict[str, Any]:
        timestamp = _parse_timestamp(payload.get("timestamp")) or datetime.utcnow().replace(microsecond=0)
        tds = int(payload.get("tds", 0))
        turbidity = float(payload.get("turbidity", 0))
        temperature = float(payload.get("temperature", 0))
        record = {
            "timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            "tds": tds,
            "turbidity": turbidity,
            "temperature": temperature,
            "status": payload.get("status") or self._quality_status(tds, turbidity, temperature),
            "device_id": payload.get("device_id", "raspberry-pi"),
        }
        if self._use_firebase():
            self._push_record("water_quality_logs", record)
            self._write_value("water_quality/latest", record)
        else:
            self.water_quality_log.insert(0, {"timestamp": timestamp, "tds": tds, "turbidity": turbidity, "temperature": temperature, "status": record["status"]})
            self.water_quality_log = self.water_quality_log[:200]
        self.last_ingest_at = datetime.utcnow()
        self.system_snapshot["last_update"] = _utc(self.last_ingest_at)
        return record

    def record_detection_sample(self, payload: dict[str, Any]) -> dict[str, Any]:
        timestamp = _parse_timestamp(payload.get("timestamp")) or datetime.utcnow().replace(microsecond=0)
        class_id = class_id_from_payload(payload)
        class_name = class_name_from_payload(payload)
        # Honor REAL per-class counts sent by the Pi detection script.
        # Fall back to class-id based single count only if not provided.
        bottle_count = int(payload.get("bottle_count", 1 if class_id == 0 else 0) or 0)
        debris_count = int(payload.get("debris_count", 1 if class_id == 1 else 0) or 0)
        total_objects = int(payload.get("total_objects", payload.get("count", 0) or 0))
        if not total_objects:
            total_objects = bottle_count + debris_count
        record = {
            "timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            "class_id": class_id,
            "class_name": class_name,
            "object_type": class_name,
            "bottle_count": bottle_count,
            "debris_count": debris_count,
            "total_objects": total_objects,
            "confidence_score": float(payload.get("confidence_score", payload.get("confidence", 0))),
            "device_id": payload.get("device_id", "raspberry-pi"),
        }
        if self._use_firebase():
            self._push_record("detection_logs", record)
            self._write_value("detection/latest", record)
        else:
            self.detection_log.insert(0, {"timestamp": timestamp, "class_id": class_id, "class_name": class_name, "object_type": class_name, "bottle_count": bottle_count, "debris_count": debris_count, "total_objects": total_objects, "confidence": record["confidence_score"]})
            self.detection_log = self.detection_log[:200]
        # ---- Cumulative counting: add only NEW objects since the last frame ----
        # (A static object staying in view is not re-counted every frame.)
        new_bottles = max(0, bottle_count - self._prev_bottle)
        new_debris = max(0, debris_count - self._prev_debris)
        self.total_bottles += new_bottles
        self.total_debris += new_debris
        self._prev_bottle = bottle_count
        self._prev_debris = debris_count
        if self._use_firebase():
            self._write_value("detection/totals", {
                "total_bottles": self.total_bottles,
                "total_debris": self.total_debris,
            })
        self.last_ingest_at = datetime.utcnow()
        self.system_snapshot["last_update"] = _utc(self.last_ingest_at)
        return record

    def record_frame(self, payload: dict[str, Any], image_file: Any | None = None) -> dict[str, Any]:
        timestamp = _parse_timestamp(payload.get("timestamp")) or datetime.utcnow().replace(microsecond=0)
        device_id = payload.get("device_id", "raspberry-pi")
        snapshot_url = payload.get("snapshot_url") or payload.get("image_url") or "/static/img/placeholder-detection.svg"

        if image_file is not None:
            raw_bytes = image_file.read()
            if raw_bytes:
                snapshot_url = f"data:{image_file.mimetype or 'image/jpeg'};base64,{base64.b64encode(raw_bytes).decode('utf-8')}"

        record = {
            "snapshot_url": snapshot_url,
            "timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            "device_id": device_id,
        }
        self.latest_frame_data = record
        self.system_snapshot["camera"] = "Streaming"
        self.system_snapshot["last_update"] = _utc(datetime.utcnow())
        return record

    # ----- User management (runtime, admin-editable) -----------------------

    VALID_ROLES = ("Admin", "Researcher", "Viewer")
    PROTECTED_EMAIL = "admin@lakewatch.local"  # cannot be deleted (prevents lockout)

    @staticmethod
    def _user_key(email: str) -> str:
        # Firebase keys can't contain . $ # [ ] / — encode the email safely.
        return base64.urlsafe_b64encode(email.strip().lower().encode()).decode().rstrip("=")

    def _find_user(self, email: str) -> UserAccount | None:
        email = email.strip().lower()
        for user in self.user_accounts:
            if user.email == email:
                return user
        return None

    def _load_persisted_users(self) -> None:
        """Merge any admin-added users stored in Firebase into the account list."""
        ref = self._root_child("users")
        if ref is None:
            return
        try:
            raw = ref.get() or {}
        except Exception as exc:  # noqa: BLE001
            print(f"[monitoring_store] Could not load users from Firebase: {exc}")
            return
        records = raw.values() if isinstance(raw, dict) else (raw if isinstance(raw, list) else [])
        for rec in records:
            if not isinstance(rec, dict) or not rec.get("email"):
                continue
            email = str(rec["email"]).strip().lower()
            account = UserAccount(
                email=email,
                password="",
                role=rec.get("role", "Viewer"),
                name=rec.get("name", email.split("@")[0].title()),
                active=bool(rec.get("active", True)),
                password_hash=rec.get("password_hash", ""),
            )
            # Replace if exists, else append.
            self.user_accounts = [u for u in self.user_accounts if u.email != email]
            self.user_accounts.append(account)

    def _persist_user(self, account: UserAccount) -> None:
        ref = self._root_child("users")
        if ref is None:
            return
        try:
            ref.child(self._user_key(account.email)).set({
                "email": account.email,
                "role": account.role,
                "name": account.name,
                "active": account.active,
                "password_hash": account.password_hash,
            })
        except Exception as exc:  # noqa: BLE001
            print(f"[monitoring_store] Could not persist user to Firebase: {exc}")

    def verify_login(self, email: str, password: str) -> str | None:
        """Return the user's role if credentials are valid, else None."""
        user = self._find_user(email)
        if user and user.active and user.password_hash and check_password_hash(user.password_hash, password):
            return user.role
        return None

    def user_exists(self, email: str) -> bool:
        return self._find_user(email) is not None

    def add_user(self, email: str, password: str, role: str, name: str = "") -> tuple[bool, str]:
        """Create or update a user. Returns (ok, message)."""
        email = (email or "").strip().lower()
        password = password or ""
        role = (role or "").strip().title()
        if "@" not in email or "." not in email:
            return False, "Please enter a valid email address."
        if len(password) < 6:
            return False, "Password must be at least 6 characters."
        if role not in self.VALID_ROLES:
            return False, f"Role must be one of: {', '.join(self.VALID_ROLES)}."

        account = UserAccount(
            email=email,
            password="",
            role=role,
            name=name.strip() or email.split("@")[0].title(),
            active=True,
            password_hash=generate_password_hash(password),
        )
        existed = self._find_user(email) is not None
        self.user_accounts = [u for u in self.user_accounts if u.email != email]
        self.user_accounts.append(account)
        self._persist_user(account)
        if not self._use_firebase():
            # No database -> the account lives only in memory and will be lost
            # on restart. Tell the admin instead of failing silently.
            return True, ("User updated, but NOT saved permanently: cloud database is off, "
                          "so this account disappears when the server restarts.")
        return True, ("User updated." if existed else "User added.")

    def delete_user(self, email: str) -> tuple[bool, str]:
        email = (email or "").strip().lower()
        if email == self.PROTECTED_EMAIL:
            return False, "The primary admin account cannot be deleted."
        if not self._find_user(email):
            return False, "User not found."
        self.user_accounts = [u for u in self.user_accounts if u.email != email]
        ref = self._root_child("users")
        if ref is not None:
            try:
                ref.child(self._user_key(email)).delete()
            except Exception as exc:  # noqa: BLE001
                print(f"[monitoring_store] Could not delete user from Firebase: {exc}")
        return True, "User removed."

    def users(self) -> list[dict[str, Any]]:
        return [
            {
                "email": user.email,
                "name": user.name,
                "role": user.role,
                "active": user.active,
            }
            for user in self.user_accounts
        ]

    def report_payload(self, filters: dict[str, str]) -> dict[str, Any]:
        detections = self.detections(filters)
        water_logs = self.water_logs(filters)

        detection_count = len(detections)
        water_count = len(water_logs)
        good_quality = sum(1 for row in water_logs if row["status"] == "Good")
        poor_quality = sum(1 for row in water_logs if row["status"] == "Poor")

        return {
            "filters": filters,
            "detections": detections,
            "water_logs": water_logs,
            "summary": {
                "detection_count": detection_count,
                "water_count": water_count,
                "good_quality": good_quality,
                "poor_quality": poor_quality,
            },
        }

    def search(self, filters: dict[str, str]) -> dict[str, list[dict[str, Any]]]:
        return {"detections": self.detections(filters), "water_logs": self.water_logs(filters)}

    def export_csv(self, filters: dict[str, str]) -> bytes:
        report = self.report_payload(filters)
        buffer = StringIO()
        writer = csv.writer(buffer)

        writer.writerow(["Smart AI Lake Cleaning Robot Monitoring System"])
        writer.writerow([f"Generated: {_utc(datetime.utcnow())}"])
        writer.writerow([])
        writer.writerow(["Summary"])
        for key, value in report["summary"].items():
            writer.writerow([key, value])
        writer.writerow([])
        writer.writerow(["Detection Logs"])
        writer.writerow(["Timestamp", "Class ID", "Class Name", "Bottle Count", "Debris Count", "Total Objects", "Confidence"])
        for row in report["detections"]:
            writer.writerow([row["timestamp"], row["class_id"], row["class_name"], row["bottle_count"], row["debris_count"], row["total_objects"], row["confidence_score"]])
        writer.writerow([])
        writer.writerow(["Water Quality Logs"])
        writer.writerow(["Timestamp", "TDS", "Turbidity", "Temperature", "Status"])
        for row in report["water_logs"]:
            writer.writerow([row["timestamp"], row["tds"], row["turbidity"], row["temperature"], row["status"]])

        return buffer.getvalue().encode("utf-8")

    def export_xlsx(self, filters: dict[str, str]) -> bytes:
        report = self.report_payload(filters)
        workbook = Workbook()
        overview_sheet = workbook.active
        overview_sheet.title = "Summary"

        header_fill = PatternFill("solid", fgColor="0F766E")
        header_font = Font(color="FFFFFF", bold=True)

        overview_sheet.append(["Smart AI Lake Cleaning Robot Monitoring System"])
        overview_sheet.append([f"Generated: {_utc(datetime.utcnow())}"])
        overview_sheet.append([])
        overview_sheet.append(["Metric", "Value"])
        for cell in overview_sheet[4]:
            cell.fill = header_fill
            cell.font = header_font
        for key, value in report["summary"].items():
            overview_sheet.append([key, value])

        detection_sheet = workbook.create_sheet("Detections")
        detection_sheet.append(["Timestamp", "Class ID", "Class Name", "Bottle Count", "Debris Count", "Total Objects", "Confidence"])
        for cell in detection_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in report["detections"]:
            detection_sheet.append([row["timestamp"], row["class_id"], row["class_name"], row["bottle_count"], row["debris_count"], row["total_objects"], row["confidence_score"]])

        water_sheet = workbook.create_sheet("Water Quality")
        water_sheet.append(["Timestamp", "TDS", "Turbidity", "Temperature", "Status"])
        for cell in water_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in report["water_logs"]:
            water_sheet.append([row["timestamp"], row["tds"], row["turbidity"], row["temperature"], row["status"]])

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def export_pdf(self, filters: dict[str, str]) -> bytes:
        report = self.report_payload(filters)
        buffer = BytesIO()
        document = SimpleDocTemplate(buffer, pagesize=landscape(letter), title="Smart AI Lake Monitoring Report")
        styles = getSampleStyleSheet()
        story: list[Any] = [Paragraph("Smart AI Lake Cleaning Robot Monitoring System", styles["Title"]), Spacer(1, 12)]
        story.append(Paragraph(f"Generated: {_utc(datetime.utcnow())}", styles["Normal"]))
        story.append(Spacer(1, 12))

        summary_data = [["Metric", "Value"]] + [[key, str(value)] for key, value in report["summary"].items()]
        summary_table = Table(summary_data, hAlign="LEFT")
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ]
            )
        )
        story.extend([Paragraph("Summary", styles["Heading2"]), summary_table, Spacer(1, 12)])

        detection_data = [["Timestamp", "Class", "Bottle", "Debris", "Total", "Confidence"]]
        detection_data.extend(
            [[row["timestamp"], f"{row['class_id']}: {row['class_name']}", str(row["bottle_count"]), str(row["debris_count"]), str(row["total_objects"]), f"{row['confidence_score']:.2f}"] for row in report["detections"]]
        )
        detection_table = Table(detection_data, repeatRows=1, hAlign="LEFT")
        detection_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#102033")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )
        story.extend([Paragraph("Detection Logs", styles["Heading2"]), detection_table, Spacer(1, 12)])

        water_data = [["Timestamp", "TDS", "Turbidity", "Temperature", "Status"]]
        water_data.extend([[row["timestamp"], str(row["tds"]), str(row["turbidity"]), str(row["temperature"]), row["status"]] for row in report["water_logs"]])
        water_table = Table(water_data, repeatRows=1, hAlign="LEFT")
        water_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )
        story.extend([Paragraph("Water Quality Logs", styles["Heading2"]), water_table])

        document.build(story)
        return buffer.getvalue()


store = MonitoringStore()
